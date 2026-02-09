import os
import uuid
from datetime import datetime
from pathlib import Path

from fastapi import FastAPI, UploadFile, File, HTTPException
from openpyxl import Workbook, load_workbook

from src.detector import run_and_save_output

app = FastAPI(title="Photo-of-Photo Detector API", version="1.0")

ROOT = Path(__file__).resolve().parent.parent
UPLOADS = ROOT / "uploads"
OUTPUTS = ROOT / "outputs"
LOGS = ROOT / "logs"
EXCEL = LOGS / "results.xlsx"

for p in (UPLOADS, OUTPUTS, LOGS):
    p.mkdir(exist_ok=True)

ALLOWED_EXT = {".jpg", ".jpeg", ".png"}


def ensure_excel():
    if EXCEL.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    ws.append(["timestamp", "input_path", "output_contoured_path", "verdict", "reason"])
    wb.save(EXCEL)


def log_row(input_path: str, output_path: str, verdict: bool, reason: str):
    ensure_excel()
    wb = load_workbook(EXCEL)
    ws = wb["results"]
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        input_path,
        output_path,
        verdict,
        reason
    ])
    wb.save(EXCEL)


@app.get("/health")
def health():
    return {"status": "ok"}


@app.post("/detect")
async def detect(file: UploadFile = File(...)):
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext and ext not in ALLOWED_EXT:
        raise HTTPException(status_code=400, detail=f"Unsupported file type: {ext}")

    uid = uuid.uuid4().hex
    input_path = UPLOADS / f"{uid}{ext or '.png'}"

    try:
        # 1) save upload
        input_path.write_bytes(await file.read())

        # 2) detector runs + saves contoured output
        verdict, info, out_path = run_and_save_output(
            str(input_path),
            output_dir=str(OUTPUTS)
        )

        reason = info.get("decision_reason", "")
        log_row(str(input_path), out_path, bool(verdict), reason)

        return {
            "verdict": bool(verdict),
            "input_path": str(input_path),
            "output_path": out_path,
            "reason": reason
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal server error: {e}")
